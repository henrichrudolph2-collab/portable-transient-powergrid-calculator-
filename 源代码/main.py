import sys
import csv
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog,
                             QTableWidgetItem, QMessageBox, QVBoxLayout)
from PyQt5.QtGui import QIcon
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from ui_PFCal import Ui_MainWindow
from Method_Callable import Methods
from help_user import Ui_Form
import resources_rc  # This loads the embedded resources


class HelpWindow(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setStyleSheet("""
            QWidget {
                border: 2px solid black;
            }
        """)
        self.setWindowTitle("用户使用指南")
        # self.textBrowser.setSource(QtCore.QUrl.fromLocalFile("help_user.htm"))
        self.textBrowser.setSource(QtCore.QUrl("qrc:/help.html"))


class MyMainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setStyleSheet("""
            QMainWindow {
                border: 2px solid black;
                background-color: rgb(116, 190, 172);
            }
        """)
        # Define expected column headers
        self.bus_keys = [
            'num', 'type', 'voltage', 'angle',
            'Pgen', 'Qgen', 'Qmin_gen', 'Qmax_gen',
            'Pload', 'Qload', 'Qshunt', 'hasGen',
            'InertiaConst', 'Damping', 'TransientReactance'
        ]
        self.line_keys = [
            'from', 'to', 'R', 'Xl',
            'Xc', 'tap', 'tapmin', 'tapmax', 'phi'
        ]
        self.setWindowTitle("一种电力系统稳态分析与电磁暂态仿真软件V1.0")
        # Initialize tables with headers
        self.actionFile.triggered.connect(self.open_help_window)
        self.init_table_headers(self.tableWidget, self.bus_keys)
        self.init_table_headers(self.tableWidget_2, self.line_keys)

        self.pushButton.clicked.connect(self.load_file_to_tableWidget1)
        self.pushButton_2.clicked.connect(self.load_file_to_tableWidget2)
        self.pushButton_3.clicked.connect(self.polar_power_flow)  # Connect compute button
        self.pushButton_4.clicked.connect(self.rectangular_axis_power_flow)  # Connect compute button
        self.pushButton_5.clicked.connect(self.compute_SBS)
        self.pushButton_6.clicked.connect(self.compute_PEBS3)
        self.pushButton_7.clicked.connect(self.save_tableWidget_3)
        self.pushButton_8.clicked.connect(self.save_tableWidget_4)

    def open_help_window(self):
        self.help_window = HelpWindow()
        self.help_window.show()

    def init_table_headers(self, table_widget, headers):
        table_widget.clear()
        table_widget.setColumnCount(len(headers))
        table_widget.setRowCount(5)  # Or any number for initial blank rows
        table_widget.setHorizontalHeaderLabels(headers)

    def load_file_to_tableWidget1(self):
        self.load_excel_to_table(self.tableWidget)

    def load_file_to_tableWidget2(self):
        self.load_excel_to_table(self.tableWidget_2)

    def load_excel_to_table(self, table_widget):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Open File", "", "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        if file_name:
            try:
                if file_name.endswith('.csv'):
                    with open(file_name, newline='', encoding='utf-8-sig') as f:
                        reader = list(csv.reader(f))
                elif file_name.endswith('.xlsx'):
                    wb = load_workbook(filename=file_name, read_only=True)
                    ws = wb.active
                    reader = [[cell.value for cell in row] for row in ws.iter_rows()]
                else:
                    return
                self.populate_table_from_array(table_widget, reader)
            except Exception as e:
                print("Error reading file:", e)

    def populate_table_from_array(self, table_widget, data):
        table_widget.clear()
        if not data or len(data) < 2:
            return

        headers = data[0]  # First row as headers
        body = data[1:]  # Remaining rows as data

        row_count = len(body)
        col_count = max(len(row) for row in body)

        table_widget.setRowCount(row_count)
        table_widget.setColumnCount(col_count)
        table_widget.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in headers])

        for i, row in enumerate(body):
            for j, val in enumerate(row):
                if val is not None:
                    table_widget.setItem(i, j, QTableWidgetItem(str(val)))

    def save_tableWidget_3(self):
        self.save_table_to_excel(self.tableWidget_3, "Save BusData", "BusData.xlsx")

    def save_tableWidget_4(self):
        self.save_table_to_excel(self.tableWidget_4, "Save LineData", "LineData.xlsx")

    def save_table_to_excel(self, table_widget, dialog_title, default_filename):
        path, _ = QFileDialog.getSaveFileName(self, dialog_title, default_filename, "Excel Files (*.xlsx)")
        if path:
            try:
                rows = table_widget.rowCount()
                cols = table_widget.columnCount()
                headers = [table_widget.horizontalHeaderItem(j).text() for j in range(cols)]

                wb = Workbook()
                ws = wb.active
                ws.append(headers)

                for i in range(rows):
                    row_data = []
                    for j in range(cols):
                        item = table_widget.item(i, j)
                        text = item.text().strip() if item else ""
                        try:
                            value = float(text)
                        except ValueError:
                            value = text
                        row_data.append(value)
                    ws.append(row_data)

                wb.save(path)
                print(f"Saved to: {path}")
            except Exception as e:
                print("Error saving file:", e)

    def polar_power_flow(self):
        try:
            # --- LOCK UI & show Busy ---
            self.setEnabled(False)
            self.statusBar().showMessage("Busy")

            # Read scalar inputs
            Sbase = float(self.lineEdit.text())
            Ubase = float(self.lineEdit_2.text())
            index = self.comboBox.currentIndex()
            if index == 2:  # third item (0-based indexing)
                f = 1 / (2*np.pi)
            else:
                selected_text = self.comboBox.currentText()
                f = float(selected_text)
            print(f"Sbase = {Sbase}, Ubase = {Ubase}, f = {f}")

            # --- Extract BusData ---
            bus_keys = [
                'num', 'type', 'voltage', 'angle',
                'Pgen', 'Qgen', 'Qmin_gen', 'Qmax_gen',
                'Pload', 'Qload', 'Qshunt', 'hasGen',
                'InertiaConst', 'Damping', 'TransientSusceptance'
            ]
            # df_bus = self.read_table_to_df(self.tableWidget)
            # if len(bus_keys) != df_bus.shape[1]:
            #     print("Mismatch in BusData columns.")
            #     return
            # BusData = {key: df_bus.iloc[:, i].tolist() for i, key in enumerate(bus_keys)}
            data_bus = self.read_table_to_array(self.tableWidget)
            BusData = {key: [row[i] for row in data_bus] for i, key in enumerate(bus_keys)}
            print("BusData loaded.")

            # --- Extract LineData ---
            line_keys = [
                'from', 'to', 'R', 'Xl',
                'Xc', 'tap', 'tapmin', 'tapmax', 'phi'
            ]
            # df_line = self.read_table_to_df(self.tableWidget_2)
            # if len(line_keys) != df_line.shape[1]:
            #     print("Mismatch in LineData columns.")
            #     return
            # LineData = {key: df_line.iloc[:, i].tolist() for i, key in enumerate(line_keys)}
            data_line = self.read_table_to_array(self.tableWidget_2)
            LineData = {key: [row[i] for row in data_line] for i, key in enumerate(line_keys)}
            print("LineData loaded.")

            BusData_1, LineData_1 = Methods.polar_power_flow(Sbase, Ubase, f, BusData, LineData)

            # --- Display BusData_1 in tableWidget_3 ---
            self.populate_table_from_dict(self.tableWidget_3, BusData_1)

            # --- Display LineData_1 in tableWidget_4 ---
            self.populate_table_from_dict(self.tableWidget_4, LineData_1)


        except Exception as e:
            print("Error in compute_variables:", e)
            # Show error message box
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Computation Error")
            msg.setText("An error occurred during power flow calculation.")
            msg.setInformativeText(str(e))
            msg.exec_()

        finally:
            self.setEnabled(True)
            self.statusBar().showMessage("Ready")

    def rectangular_axis_power_flow(self):
        try:
            # --- LOCK UI & show Busy ---
            self.setEnabled(False)
            self.statusBar().showMessage("Busy")

            # Read scalar inputs
            Sbase = float(self.lineEdit.text())
            Ubase = float(self.lineEdit_2.text())
            index = self.comboBox.currentIndex()
            if index == 2:  # third item (0-based indexing)
                f = 1 / (2*np.pi)
            else:
                selected_text = self.comboBox.currentText()
                f = float(selected_text)
            print(f"Sbase = {Sbase}, Ubase = {Ubase}, f = {f}")

            # --- Extract BusData ---
            bus_keys = [
                'num', 'type', 'voltage', 'angle',
                'Pgen', 'Qgen', 'Qmin_gen', 'Qmax_gen',
                'Pload', 'Qload', 'Qshunt', 'hasGen',
                'InertiaConst', 'Damping', 'TransientSusceptance'
            ]
            # df_bus = self.read_table_to_df(self.tableWidget)
            # if len(bus_keys) != df_bus.shape[1]:
            #     print("Mismatch in BusData columns.")
            #     return
            # BusData = {key: df_bus.iloc[:, i].tolist() for i, key in enumerate(bus_keys)}
            data_bus = self.read_table_to_array(self.tableWidget)
            BusData = {key: [row[i] for row in data_bus] for i, key in enumerate(bus_keys)}
            print("BusData loaded.")

            # --- Extract LineData ---
            line_keys = [
                'from', 'to', 'R', 'Xl',
                'Xc', 'tap', 'tapmin', 'tapmax', 'phi'
            ]
            # df_line = self.read_table_to_df(self.tableWidget_2)
            # if len(line_keys) != df_line.shape[1]:
            #     print("Mismatch in LineData columns.")
            #     return
            # LineData = {key: df_line.iloc[:, i].tolist() for i, key in enumerate(line_keys)}
            data_line = self.read_table_to_array(self.tableWidget_2)
            LineData = {key: [row[i] for row in data_line] for i, key in enumerate(line_keys)}
            print("LineData loaded.")

            BusData_1, LineData_1 = Methods.rec_power_flow(Sbase, Ubase, f, BusData, LineData)

            # --- Display BusData_1 in tableWidget_3 ---
            self.populate_table_from_dict(self.tableWidget_3, BusData_1)

            # --- Display LineData_1 in tableWidget_4 ---
            self.populate_table_from_dict(self.tableWidget_4, LineData_1)


        except Exception as e:
            print("Error in compute_variables:", e)
            # Show error message box
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Computation Error")
            msg.setText("An error occurred during power flow calculation.")
            msg.setInformativeText(str(e))
            msg.exec_()

        finally:
            self.setEnabled(True)
            self.statusBar().showMessage("Ready")

    def compute_SBS(self):
        try:
            # --- LOCK UI & show Busy ---
            self.setEnabled(False)
            self.statusBar().showMessage("Busy...")

            # Read scalar inputs
            Sbase = float(self.lineEdit.text())
            Ubase = float(self.lineEdit_2.text())
            index = self.comboBox.currentIndex()
            if index == 2:  # third item (0-based indexing)
                f = 1 / (2*np.pi)
            else:
                selected_text = self.comboBox.currentText()
                f = float(selected_text)
            ContingencyBus = int(self.lineEdit_4.text())  # Indicates the bus where the contingency occurs
            line = int(self.lineEdit_5.text())  # Defines which line will be removed after the contingency
            from_bus = 1
            to_bus = int(self.lineEdit_7.text())
            tfout0 = float(self.lineEdit_8.text())  # Initial fault time
            tfoult = float(self.lineEdit_9.text())  # Fault clearing time (in seconds)
            tAfter = float(self.lineEdit_10.text())  # Maximum post-fault simulation time
            maxstep = float(self.lineEdit_11.text())  # Integration time step

            print(f"Sbase = {Sbase}, Ubase = {Ubase}, f = {f}")
            epochs = 1

            if maxstep <= 0.001:
                maxstep = 0.001

            # --- Extract BusData ---
            bus_keys_1 = [
                'num', 'type', 'voltage', 'angle',
                'angle_radian', 'Pgen', 'Qgen', 'Qmin_gen', 'Qmax_gen',
                'Pload', 'Qload', 'Qshunt', 'hasGen',
                'InertiaConst', 'Damping', 'TransientSusceptance',
                'Pk', 'Qk'
            ]
            # df_bus = self.read_table_to_df(self.tableWidget_3)
            # if len(bus_keys_1) != df_bus.shape[1]:
            #     print("Mismatch in BusData columns.")
            #     return
            # BusData = {key: df_bus.iloc[:, i].tolist() for i, key in enumerate(bus_keys_1)}
            data_bus = self.read_table_to_array(self.tableWidget_3)
            BusData = {key: [row[i] for row in data_bus] for i, key in enumerate(bus_keys_1)}
            print("BusData loaded.")

            # --- Extract LineData ---
            line_keys_1 = [
                'from', 'to', 'R', 'Xl',
                'Xc', 'tap', 'tapmin', 'tapmax', 'phi',
                'Pkm', 'Pmk', 'Qkm', 'Qmk', 'ActiveLoss_km', 'ReactiveLoss_km'
            ]
            # df_line = self.read_table_to_df(self.tableWidget_4)
            # if len(line_keys_1) != df_line.shape[1]:
            #     print("Mismatch in LineData columns.")
            #     return
            # LineData = {key: df_line.iloc[:, i].tolist() for i, key in enumerate(line_keys_1)}
            data_line = self.read_table_to_array(self.tableWidget_4)
            LineData = {key: [row[i] for row in data_line] for i, key in enumerate(line_keys_1)}
            print("LineData loaded.")

            # fig = Transient.PEBS3(Sbase, Ubase, f, BusData, LineData)
            # self.display_plot_in_tab4(fig)
            # fig.savefig('C:/Users/henri/Desktop/PEBS3_[%d]sw1.png' % epochs)

            fig1, fig2 = Methods.step_by_step(Sbase, Ubase, f, BusData, LineData, ContingencyBus, line,
                                              from_bus, to_bus, tfout0, tfoult, tAfter, maxstep)
            self.display_plot_in_tab3_4(fig1, fig2)

        except Exception as e:
            # print("Error in compute_variables:", e)
            # Show error message box
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Computation Error")
            msg.setText("An error occurred during Transient simulation Step-By-Step.")
            msg.setInformativeText(str(e))
            msg.exec_()

        finally:
            # --- UNLOCK UI & show Ready ---
            self.setEnabled(True)
            self.statusBar().showMessage("Ready")

    def compute_PEBS3(self):
        try:
            # --- LOCK UI & show Busy ---
            self.setEnabled(False)
            self.statusBar().showMessage("Busy")

            # Read scalar inputs
            Sbase = float(self.lineEdit.text())
            Ubase = float(self.lineEdit_2.text())
            index = self.comboBox.currentIndex()
            if index == 2:  # third item (0-based indexing)
                f = 1 / (2*np.pi)
            else:
                selected_text = self.comboBox.currentText()
                f = float(selected_text)
            ContingencyBus = int(self.lineEdit_4.text())  # Indicates the bus where the contingency occurs
            line = int(self.lineEdit_5.text())  # Defines which line will be removed after the contingency
            from_bus = 1
            to_bus = int(self.lineEdit_7.text())
            tfout0 = float(self.lineEdit_8.text())  # Initial fault time
            tfoult = float(self.lineEdit_9.text())  # Fault clearing time (in seconds)
            tAfter = float(self.lineEdit_10.text())  # Maximum post-fault simulation time
            maxstep = float(self.lineEdit_11.text())  # Integration time step

            print(f"Sbase = {Sbase}, Ubase = {Ubase}, f = {f}")
            epochs = 1

            if maxstep <= 0.001:
                maxstep = 0.001

            if tfoult <= tfout0 + 0.8:
                tfoult = tfout0 + 0.8

            # --- Extract BusData ---
            bus_keys_1 = [
                'num', 'type', 'voltage', 'angle',
                'angle_radian', 'Pgen', 'Qgen', 'Qmin_gen', 'Qmax_gen',
                'Pload', 'Qload', 'Qshunt', 'hasGen',
                'InertiaConst', 'Damping', 'TransientSusceptance',
                'Pk', 'Qk'
            ]
            data_bus = self.read_table_to_array(self.tableWidget_3)
            BusData = {key: [row[i] for row in data_bus] for i, key in enumerate(bus_keys_1)}
            print("BusData loaded.")

            # --- Extract LineData ---
            line_keys_1 = [
                'from', 'to', 'R', 'Xl',
                'Xc', 'tap', 'tapmin', 'tapmax', 'phi',
                'Pkm', 'Pmk', 'Qkm', 'Qmk', 'ActiveLoss_km', 'ReactiveLoss_km'
            ]
            data_line = self.read_table_to_array(self.tableWidget_4)
            LineData = {key: [row[i] for row in data_line] for i, key in enumerate(line_keys_1)}
            print("LineData loaded.")

            # plt = Transient.PEBS3(Sbase, Ubase, f, BusData, LineData)
            # plt.savefig('C:/Users/henri/Desktop/PEBS3_[%d]sw1.png'% epochs)
            # plt.show()

            fig = Methods.pebs3(Sbase, Ubase, f, BusData, LineData, ContingencyBus, line,
                                from_bus, to_bus, tfout0, tfoult, tAfter, maxstep)

            # Transient.PEBS3(Sbase, Ubase, f, BusData, LineData)
            # fig = plt.gcf()  # Get the current figure object

            self.display_plot_in_tab5(fig)
            # fig.savefig('C:/Users/henri/Desktop/PEBS3_[%d]sw1.png' % epochs)

        except Exception as e:
            print("Error in compute_variables:", e)
            # Show error message box
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Computation Error")
            msg.setText("An error occurred during Transient simulation PEBS.")
            msg.setInformativeText(str(e))
            msg.exec_()

        finally:
            self.setEnabled(True)
            self.statusBar().showMessage("Ready")

    def read_table_to_array(self, table_widget):
        rows = table_widget.rowCount()
        cols = table_widget.columnCount()
        data = []

        for i in range(rows):
            row_data = []
            for j in range(cols):
                item = table_widget.item(i, j)
                text = item.text().strip() if item else ''
                try:
                    value = float(text)
                except ValueError:
                    value = 0.0  # or np.nan if preferred
                row_data.append(value)
            data.append(row_data)
        return data

    def populate_table_from_dict(self, table_widget, data_dict):
        """
        Populates a QTableWidget from a dictionary of lists.
        Each key becomes a column header.
        """
        if not data_dict:
            return

        headers = list(data_dict.keys())
        num_rows = len(next(iter(data_dict.values())))
        num_cols = len(headers)

        table_widget.clear()
        table_widget.setRowCount(num_rows)
        table_widget.setColumnCount(num_cols)
        table_widget.setHorizontalHeaderLabels(headers)

        for col, key in enumerate(headers):
            for row, value in enumerate(data_dict[key]):
                item = QTableWidgetItem(str(value))
                table_widget.setItem(row, col, item)

    def display_plot_in_tab3_4(self, fig1, fig2):
        """Embed two matplotlib figures into widget and widget_2 in tab_4."""

        # --- Handle widget 1 ---
        layout1 = self.tab_3.layout()
        if layout1 is None:
            layout1 = QVBoxLayout()
            self.tab_3.setLayout(layout1)
        else:
            while layout1.count():
                child = layout1.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()

        canvas1 = FigureCanvas(fig1)
        layout1.addWidget(canvas1)
        canvas1.draw()

        # --- Handle widget 2 ---
        layout2 = self.tab_4.layout()
        if layout2 is None:
            layout2 = QVBoxLayout()
            self.tab_4.setLayout(layout2)
        else:
            while layout2.count():
                child = layout2.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()

        canvas2 = FigureCanvas(fig2)
        layout2.addWidget(canvas2)
        canvas2.draw()

    def display_plot_in_tab5(self, fig):
        """Embed a matplotlib figure into tab_3."""
        # Remove any existing widgets from tab_3
        layout = self.tab_5.layout()
        if layout is None:
            layout = QVBoxLayout()
            self.tab_5.setLayout(layout)
        else:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()

        # Create canvas and add to layout
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)
        canvas.draw()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    # app.setWindowIcon(QIcon('001.jpg'))
    app.setWindowIcon(QIcon(':/window_icon.jpg'))
    window = MyMainWindow()
    window.show()
    sys.exit(app.exec_())
